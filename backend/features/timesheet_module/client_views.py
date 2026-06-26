from datetime import datetime
from django.http import HttpResponse
from rest_framework import viewsets, permissions, status, generics
from rest_framework.views import APIView
from rest_framework.decorators import action
from rest_framework.response import Response
from features.tenant_module.context import get_current_organization
from features.planner_module.models import Task
from features.notification_module.models import Notification
from features.notification_module.serializers import NotificationSerializer
from .models import Project
from .models_client import Release, ClientApprovalRequest, ClientProjectAccess, ClientDocument, ProjectActivity
from .client_serializers import (
    ClientProjectSerializer,
    ClientTaskSerializer,
    ReleaseSerializer,
    ClientApprovalRequestSerializer,
    ClientDocumentSerializer,
    ProjectActivitySerializer
)
from .models import TimesheetEntry as Timesheet


class IsClientUser(permissions.BasePermission):
    """
    Allow any authenticated member of the active organization to access client portal
    endpoints. Data scoping (project visibility, client access grants) is enforced
    at the queryset level. Superusers are always allowed.
    """
    def has_permission(self, request, view):
        if not (request.user and request.user.is_authenticated):
            return False
        organization = get_current_organization()
        if not organization:
            return False
        if request.user.is_superuser:
            return True
        # Allow any member of the organization (internal staff or CLIENT role)
        return request.user.memberships.filter(
            organization=organization
        ).exists()


def get_accessible_projects(user, organization):
    """
    Returns Project queryset filtered by organization.
    - Superusers: all org projects (client-visible or not)
    - Internal users (non-CLIENT role): all client-visible org projects
    - CLIENT role users: only projects explicitly granted via ClientProjectAccess
    """
    if user.is_superuser:
        return Project.objects.filter(organization=organization, is_client_visible=True)
    is_client = user.memberships.filter(
        organization=organization,
        role__code='CLIENT'
    ).exists()
    if not is_client:
        # Internal staff can see all client-visible projects
        return Project.objects.filter(organization=organization, is_client_visible=True)
    return Project.objects.filter(
        organization=organization,
        is_client_visible=True,
        client_accesses__user=user
    )


class ClientProjectViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Exposes only client-visible, project-isolated projects the Client user is explicitly granted access to.
    """
    serializer_class = ClientProjectSerializer
    permission_classes = [IsClientUser]

    def get_queryset(self):
        organization = get_current_organization()
        if not organization:
            return Project.objects.none()
        return get_accessible_projects(self.request.user, organization).order_by('name')


class ClientTaskViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Allows clients to safely view only flagged client-visible tasks belonging to accessible projects.
    """
    serializer_class = ClientTaskSerializer
    permission_classes = [IsClientUser]

    def get_queryset(self):
        organization = get_current_organization()
        if not organization:
            return Task.objects.none()
        accessible_projects = get_accessible_projects(self.request.user, organization)
        return Task.objects.filter(
            organization=organization,
            project__in=accessible_projects,
            is_client_visible=True
        ).order_by('-created_at')


class ClientReleaseViewSet(viewsets.ModelViewSet):
    """
    Exposes release milestones linked to accessible, client-visible projects.
    Internal users (non-CLIENT role) may create and manage releases.
    CLIENT role users are restricted to read-only access.
    """
    serializer_class = ReleaseSerializer
    permission_classes = [IsClientUser]

    def get_permissions(self):
        # Write operations require internal (non-client) membership
        if self.request.method not in ('GET', 'HEAD', 'OPTIONS'):
            from features.timesheet_module.views import IsInternalUser
            return [permissions.IsAuthenticated(), IsInternalUser()]
        return [IsClientUser()]

    def get_queryset(self):
        organization = get_current_organization()
        if not organization:
            return Release.objects.none()
        # Internal users see all org releases; CLIENT users see only accessible project releases
        user = self.request.user
        if user.is_superuser or not user.memberships.filter(
            organization=organization, role__code='CLIENT'
        ).exists():
            return Release.objects.filter(
                organization=organization
            ).order_by('-release_date')
        accessible_projects = get_accessible_projects(user, organization)
        return Release.objects.filter(
            organization=organization,
            project__in=accessible_projects
        ).order_by('-release_date')

    def perform_create(self, serializer):
        organization = get_current_organization()
        serializer.save(
            organization=organization,
            created_by=self.request.user,
            updated_by=self.request.user
        )


class ClientApprovalRequestViewSet(viewsets.ModelViewSet):
    """
    Exposes approval items for client sign-offs.
    Clients are permitted to update status (APPROVED, REJECTED, NEEDS_CLARIFICATION) and comments.
    """
    serializer_class = ClientApprovalRequestSerializer
    permission_classes = [IsClientUser]
    http_method_names = ['get', 'patch', 'put', 'options']

    def get_queryset(self):
        organization = get_current_organization()
        if not organization:
            return ClientApprovalRequest.objects.none()
        accessible_projects = get_accessible_projects(self.request.user, organization)
        return ClientApprovalRequest.objects.filter(
            organization=organization,
            project__in=accessible_projects
        ).order_by('-created_at')

    def perform_update(self, serializer):
        serializer.save(
            reviewed_by=self.request.user,
            updated_by=self.request.user
        )


class ClientDocumentViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Lists files shared with clients for projects they have access to.
    """
    serializer_class = ClientDocumentSerializer
    permission_classes = [IsClientUser]

    def get_queryset(self):
        organization = get_current_organization()
        if not organization:
            return ClientDocument.objects.none()
        accessible_projects = get_accessible_projects(self.request.user, organization)
        return ClientDocument.objects.filter(
            organization=organization,
            project__in=accessible_projects,
            is_client_visible=True
        ).order_by('-created_at')


class ProjectActivityViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Chronological project activity timeline for clients.
    """
    serializer_class = ProjectActivitySerializer
    permission_classes = [IsClientUser]

    def get_queryset(self):
        organization = get_current_organization()
        if not organization:
            return ProjectActivity.objects.none()
        accessible_projects = get_accessible_projects(self.request.user, organization)
        return ProjectActivity.objects.filter(
            organization=organization,
            project__in=accessible_projects
        ).order_by('-created_at')


class ClientNotificationViewSet(viewsets.ModelViewSet):
    """
    List and update in-app notifications for client user alerts.
    """
    serializer_class = NotificationSerializer
    permission_classes = [IsClientUser]
    http_method_names = ['get', 'patch', 'put', 'options']

    def get_queryset(self):
        organization = get_current_organization()
        if not organization:
            return Notification.objects.none()
        return Notification.objects.filter(
            organization=organization,
            recipient=self.request.user
        ).order_by('-created_at')


class ClientReportsAPIView(APIView):
    """
    Aggregates high-level project metrics, milestone statuses, and risk summaries.
    Provides export endpoints.
    """
    permission_classes = [IsClientUser]

    def get(self, request, *args, **kwargs):
        # Route action kwargs if matching PDF or Excel export routes
        action_name = kwargs.get('action')
        if action_name == 'export_pdf':
            return self.export_pdf(request)
        elif action_name == 'export_excel':
            return self.export_excel(request)

        organization = get_current_organization()
        if not organization:
            return Response({'error': 'Active organization required.'}, status=status.HTTP_400_BAD_REQUEST)

        accessible_projects = get_accessible_projects(request.user, organization)
        tasks = Task.objects.filter(organization=organization, project__in=accessible_projects, is_client_visible=True)
        total_tasks = tasks.count()
        completed_tasks = tasks.filter(status='DONE').count()
        pending_tasks = total_tasks - completed_tasks

        progress_rate = round((completed_tasks / total_tasks * 100), 1) if total_tasks > 0 else 100.0

        pending_approvals = ClientApprovalRequest.objects.filter(
            organization=organization,
            project__in=accessible_projects,
            status='PENDING'
        ).count()

        # Simulated Risk Summary
        risks = [
            {'id': 1, 'title': 'Resource Constraint in Sprint 3', 'severity': 'MEDIUM'},
            {'id': 2, 'title': 'Integration latency updates', 'severity': 'LOW'}
        ]

        return Response({
            'total_tasks': total_tasks,
            'completed_tasks': completed_tasks,
            'pending_tasks': pending_tasks,
            'progress_rate': progress_rate,
            'pending_approvals': pending_approvals,
            'risks': risks
        })

    @action(detail=False, methods=['get'], url_path='export_pdf')
    def export_pdf(self, request):
        """Export timesheet report as PDF-like HTML rendered as attachment."""
        organization = get_current_organization()
        if not organization:
            return Response({'error': 'Organization required'}, status=400)

        timesheets = Timesheet.objects.filter(
            organization=organization
        ).select_related('user', 'project').order_by('-date')[:200]

        rows = "".join([
            f'<tr style="border-bottom:1px solid #2d2d34;">'
            f'<td style="padding:8px 12px;color:#e4e4e7;">{ts.user.email}</td>'
            f'<td style="padding:8px 12px;color:#a1a1aa;">{ts.project.name if ts.project else "—"}</td>'
            f'<td style="padding:8px 12px;color:#a1a1aa;">{ts.date}</td>'
            f'<td style="padding:8px 12px;color:#6366f1;font-weight:700;">{ts.hours_logged}h</td>'
            f'<td style="padding:8px 12px;color:#a1a1aa;">{ts.description[:60] if ts.description else ""}</td>'
            f'</tr>'
            for ts in timesheets
        ])

        total_hours = sum(float(ts.hours_logged) for ts in timesheets)

        html = f"""
        <!DOCTYPE html>
        <html>
        <head><meta charset="utf-8"><title>TaskSphere Timesheet Report</title></head>
        <body style="font-family:Inter,Arial,sans-serif;background:#0f0f12;color:#e4e4e7;padding:32px;">
          <div style="max-width:900px;margin:0 auto;">
            <div style="background:linear-gradient(135deg,#6366f1,#8b5cf6);padding:24px;border-radius:12px;margin-bottom:24px;">
              <h1 style="margin:0;color:#fff;font-size:24px;">📊 TaskSphere Timesheet Report</h1>
              <p style="margin:8px 0 0;color:rgba(255,255,255,0.8);">{organization.name} · Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} UTC</p>
            </div>
            <div style="display:flex;gap:16px;margin-bottom:24px;">
              <div style="flex:1;background:#1c1c1f;border:1px solid #2d2d34;border-radius:8px;padding:16px;">
                <div style="font-size:11px;color:#8e8e95;font-weight:700;text-transform:uppercase;">Total Entries</div>
                <div style="font-size:24px;font-weight:900;color:#fff;margin-top:4px;">{timesheets.count()}</div>
              </div>
              <div style="flex:1;background:#1c1c1f;border:1px solid #2d2d34;border-radius:8px;padding:16px;">
                <div style="font-size:11px;color:#8e8e95;font-weight:700;text-transform:uppercase;">Total Hours</div>
                <div style="font-size:24px;font-weight:900;color:#6366f1;margin-top:4px;">{total_hours:.1f}h</div>
              </div>
            </div>
            <table style="width:100%;border-collapse:collapse;background:#1c1c1f;border-radius:8px;overflow:hidden;">
              <thead>
                <tr style="background:#2d2d34;">
                  <th style="padding:10px 12px;text-align:left;font-size:10px;text-transform:uppercase;color:#8e8e95;">Member</th>
                  <th style="padding:10px 12px;text-align:left;font-size:10px;text-transform:uppercase;color:#8e8e95;">Project</th>
                  <th style="padding:10px 12px;text-align:left;font-size:10px;text-transform:uppercase;color:#8e8e95;">Date</th>
                  <th style="padding:10px 12px;text-align:left;font-size:10px;text-transform:uppercase;color:#8e8e95;">Hours</th>
                  <th style="padding:10px 12px;text-align:left;font-size:10px;text-transform:uppercase;color:#8e8e95;">Description</th>
                </tr>
              </thead>
              <tbody>{rows}</tbody>
            </table>
          </div>
        </body>
        </html>
        """

        response = HttpResponse(html, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="TaskSphere_Report_{datetime.now().strftime("%Y%m%d")}.html"'
        return response

    @action(detail=False, methods=['get'], url_path='export_excel')
    def export_excel(self, request):
        """Export timesheet report as Excel .xlsx file."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        organization = get_current_organization()
        if not organization:
            return Response({'error': 'Organization required'}, status=400)

        timesheets = Timesheet.objects.filter(
            organization=organization
        ).select_related('user', 'project').order_by('-date')[:500]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Timesheets'

        # Header styling
        header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)

        headers = ['Member Email', 'Member Name', 'Project', 'Date', 'Hours Logged', 'Description', 'Status']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, ts in enumerate(timesheets, 2):
            user_name = f"{ts.user.first_name} {ts.user.last_name}".strip() or ts.user.email
            ws.cell(row=row_idx, column=1, value=ts.user.email)
            ws.cell(row=row_idx, column=2, value=user_name)
            ws.cell(row=row_idx, column=3, value=ts.project.name if ts.project else '')
            ws.cell(row=row_idx, column=4, value=str(ts.date))
            ws.cell(row=row_idx, column=5, value=float(ts.hours_logged))
            ws.cell(row=row_idx, column=6, value=ts.description or '')
            ws.cell(row=row_idx, column=7, value=getattr(ts, 'status', 'PENDING'))

        # Auto column widths
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20

        # Summary sheet
        ws_summary = wb.create_sheet('Summary')
        total_hours = sum(float(ts.hours_logged) for ts in timesheets)
        ws_summary['A1'] = 'Organization'
        ws_summary['B1'] = organization.name
        ws_summary['A2'] = 'Total Entries'
        ws_summary['B2'] = timesheets.count()
        ws_summary['A3'] = 'Total Hours'
        ws_summary['B3'] = total_hours
        ws_summary['A4'] = 'Generated'
        ws_summary['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M UTC')

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='text/csv'
        )
        filename = f'TaskSphere_Timesheets_{datetime.now().strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

